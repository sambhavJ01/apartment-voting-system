"""
Report service — per-topic reporting, CSV/Excel export, audit log retrieval.
"""
import csv
import io
import logging
from datetime import datetime
from typing import List, Optional

from sqlalchemy import and_
from sqlalchemy.orm import Session

from backend.models.audit import AuditLog
from backend.models.topic import Topic
from backend.models.user import User, UserStatus
from backend.models.vote import Vote, VoteTracking

logger = logging.getLogger(__name__)


class ReportService:
    def __init__(self, db: Session) -> None:
        self.db = db

    # ── Per-topic statistics ─────────────────────────────────────────────────

    def get_topic_report(self, topic_id: int) -> dict:
        topic = self.db.query(Topic).filter(Topic.id == topic_id).first()
        if not topic:
            return {"success": False, "message": "Topic not found."}

        eligible = (
            self.db.query(User)
            .filter(
                User.status == UserStatus.APPROVED,
                User.is_active.is_(True),
                User.is_admin.is_(False),
            )
            .count()
        )
        total_votes = (
            self.db.query(VoteTracking)
            .filter(VoteTracking.topic_id == topic_id)
            .count()
        )
        participation = round((total_votes / eligible * 100) if eligible else 0, 1)

        options_stats = []
        for opt in sorted(topic.options, key=lambda o: o.order):
            count = (
                self.db.query(Vote)
                .filter(and_(Vote.topic_id == topic_id, Vote.option_id == opt.id))
                .count()
            )
            pct = round((count / total_votes * 100) if total_votes else 0, 1)
            options_stats.append(
                {
                    "option_id": opt.id,
                    "option_text": opt.text,
                    "vote_count": count,
                    "percentage": pct,
                }
            )

        return {
            "success": True,
            "topic_id": topic_id,
            "title": topic.title,
            "description": topic.description,
            "mode": topic.mode.value,
            "status": topic.status.value,
            "start_time": topic.start_time.isoformat() if topic.start_time else None,
            "end_time": topic.end_time.isoformat() if topic.end_time else None,
            "total_eligible": eligible,
            "total_votes": total_votes,
            "participation_pct": participation,
            "options": options_stats,
        }

    # ── All-topics summary ───────────────────────────────────────────────────

    def get_all_topics_summary(self) -> List[dict]:
        topics = (
            self.db.query(Topic).order_by(Topic.created_at.desc()).all()
        )
        eligible = (
            self.db.query(User)
            .filter(
                User.status == UserStatus.APPROVED,
                User.is_active.is_(True),
                User.is_admin.is_(False),
            )
            .count()
        )
        result = []
        for t in topics:
            votes = (
                self.db.query(VoteTracking)
                .filter(VoteTracking.topic_id == t.id)
                .count()
            )
            result.append(
                {
                    "topic_id": t.id,
                    "title": t.title,
                    "mode": t.mode.value,
                    "status": t.status.value,
                    "total_votes": votes,
                    "total_eligible": eligible,
                    "participation_pct": round(
                        (votes / eligible * 100) if eligible else 0, 1
                    ),
                }
            )
        return result

    # ── CSV export ───────────────────────────────────────────────────────────

    def export_csv(self, topic_id: int) -> Optional[bytes]:
        report = self.get_topic_report(topic_id)
        if not report.get("success"):
            return None

        output = io.StringIO()
        w = csv.writer(output)
        w.writerow(["Apartment Voting System — Results Report"])
        w.writerow(["Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")])
        w.writerow([])
        w.writerow(["Topic", report["title"]])
        w.writerow(["Description", report.get("description") or ""])
        w.writerow(["Mode", report["mode"].capitalize()])
        w.writerow(["Status", report["status"].capitalize()])
        w.writerow(["Total Eligible Voters", report["total_eligible"]])
        w.writerow(["Total Votes Cast", report["total_votes"]])
        w.writerow(["Participation %", f"{report['participation_pct']}%"])
        w.writerow([])
        w.writerow(["Option", "Votes", "Percentage"])
        for opt in report["options"]:
            w.writerow([opt["option_text"], opt["vote_count"], f"{opt['percentage']}%"])
        return output.getvalue().encode("utf-8")

    # ── Excel export ─────────────────────────────────────────────────────────

    def export_excel(self, topic_id: int) -> Optional[bytes]:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl not installed — cannot export Excel")
            return None

        report = self.get_topic_report(topic_id)
        if not report.get("success"):
            return None

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"

        # ── Header ────────────────────────────────────────────────────────────
        ws.merge_cells("A1:D1")
        ws["A1"] = "Apartment Voting System — Results Report"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        meta = [
            ("Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
            ("Topic", report["title"]),
            ("Description", report.get("description") or ""),
            ("Mode", report["mode"].capitalize()),
            ("Status", report["status"].capitalize()),
            ("Total Eligible Voters", report["total_eligible"]),
            ("Total Votes Cast", report["total_votes"]),
            ("Participation %", f"{report['participation_pct']}%"),
        ]
        for row_idx, (label, value) in enumerate(meta, start=2):
            ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)

        # ── Option table ──────────────────────────────────────────────────────
        header_row = len(meta) + 3
        blue_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        for col, heading in enumerate(["Option", "Votes", "Percentage"], start=1):
            cell = ws.cell(row=header_row, column=col, value=heading)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = blue_fill

        for data_row, opt in enumerate(report["options"], start=header_row + 1):
            ws.cell(row=data_row, column=1, value=opt["option_text"])
            ws.cell(row=data_row, column=2, value=opt["vote_count"])
            ws.cell(row=data_row, column=3, value=f"{opt['percentage']}%")

        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                max_len + 4, 60
            )

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── Audit log retrieval ──────────────────────────────────────────────────

    def get_audit_logs(
        self,
        limit: int = 200,
        action_filter: Optional[str] = None,
    ) -> List[dict]:
        q = self.db.query(AuditLog).order_by(AuditLog.timestamp.desc())
        if action_filter:
            q = q.filter(AuditLog.action.ilike(f"%{action_filter}%"))
        logs = q.limit(limit).all()
        return [
            {
                "id": log.id,
                "user_id": log.user_id,
                "apartment_id": log.apartment_id,
                "action": log.action,
                "timestamp": log.timestamp.isoformat(),
                "metadata": log.log_data,
                "ip_address": log.ip_address,
            }
            for log in logs
        ]
